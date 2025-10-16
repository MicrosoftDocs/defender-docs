#!/usr/bin/env python3
"""
Simple XLS/XLSX -> Markdown table converter.

Reads a sheet, uses the first non-empty row as headers, and outputs a markdown table.
"""
import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except Exception:
    openpyxl = None

try:
    import xlrd
except Exception:
    xlrd = None


def read_xlsx(path, sheet_name=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(["" if c is None else str(c) for c in row])
    return rows


def read_xls(path, sheet_name=None):
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_name(sheet_name) if sheet_name else book.sheet_by_index(0)
    rows = []
    for r in range(sheet.nrows):
        row = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
        rows.append(row)
    return rows


def rows_to_md_table(rows):
    if not rows:
        return ""
    # find first non-empty row
    header_idx = 0
    for i, r in enumerate(rows):
        if any(cell.strip() for cell in r if cell is not None):
            header_idx = i
            break
    headers = [h.strip() for h in rows[header_idx]]
    data_rows = rows[header_idx+1:]

    # normalize column count
    max_cols = max(len(headers), max((len(r) for r in data_rows), default=0))
    headers = (headers + [""] * max_cols)[:max_cols]

    md = []
    md.append("|" + "|".join(headers) + "|")
    md.append("|" + "|".join(["-" * max(1, len(h)) for h in headers]) + "|")
    for r in data_rows:
        cells = [(c.strip() if c is not None else "") for c in r]
        cells = (cells + [""] * max_cols)[:max_cols]
        md.append("|" + "|".join(cells) + "|")
    return "\n".join(md)


def build_page(title, table_md):
    header = f"---\ntitle: {title}\n---\n\n"
    content = f"# {title}\n\n{table_md}\n"
    return header + content


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("input", help="Input xls/xlsx file")
    parser.add_argument("output", help="Output markdown file")
    parser.add_argument("--sheet", help="Sheet name to read", default=None)
    parser.add_argument("--title", help="Page title", default="Converted from XLS")
    args = parser.parse_args()

    p = Path(args.input)
    if not p.exists():
        print(f"File not found: {p}")
        sys.exit(2)

    if p.suffix.lower() in (".xlsx",):
        if openpyxl is None:
            print("openpyxl is required to read .xlsx files. Install with pip install openpyxl")
            sys.exit(2)
        rows = read_xlsx(p, args.sheet)
    else:
        if xlrd is None:
            print("xlrd is required to read .xls files. Install with pip install xlrd")
            sys.exit(2)
        rows = read_xls(p, args.sheet)

    table_md = rows_to_md_table(rows)
    page = build_page(args.title, table_md)

    out = Path(args.output)
    out.write_text(page, encoding="utf-8")
    print(f"Wrote {out}")


if __name__ == '__main__':
    main()
